import argparse
import os

import pandas as pd
from openpyxl import load_workbook
import datetime

desired_width = 179

'''
Header Fields
Listing Number,Status,Property Sub-Type,Street #,Street Name,
Original List Price,Listing Price,Sold Price,Cumulative DOM,
Days On Market,# Bedrooms,Baths - Total,Baths - Full,Baths - 3/4,
Baths - 1/2,Baths - 1/4,Approx SqFt,Lot SqFt,Year Built,Sold Date
'''

def do_convert(file_name):
    # file = 'latestShoresSoldSfrYrToDate.xlsx' if file_name is None else file_name
    file = 'shores.xlsx' if file_name is None else file_name
    root, ext = os.path.splitext(file)
    if ext == '.csv':
        df = pd.read_csv(file)
    elif ext == '.xlsx':
        df = pd.read_excel(file)
    else:
        return None

    Bath = []
    for row in df:
        df['Baths - Total'] = df['Baths - Full'] + df['Baths - 3/4'] * .75 + \
                              df['Baths - 1/2'] * .5 + df['Baths - 1/4'] * .25
    if args.verbose: print(df.head())

    df.drop(['MLS\xa0#', 'Status', 'BA F', 'BA 3/4', 'BA 1/2', 'BA 1/4'],
            axis=1,
            inplace=True)
    df.drop(df.columns[[0]], axis=1)
    df.rename(columns={'BD': 'Bath', 'T BA': 'Bath', 'SqFt': 'Interior Size',
                       'Yr Blt': 'Yr. Built', 'Lot SqFt': 'Lot Size',
                       'Price': 'Sale Price', 'Address': 'Sold Listing',
                       'CDOM': 'Days on Market'},
              inplace=True)
    cols = df.columns.tolist()
    # cols

    if args.verbose: print(df.head())

    df['Sold Date'] = df['Sold Date'].apply(lambda x:
                                            '' if pd.isnull(x) else x.strftime(
                                                '%m/%d/%y'))
    if args.verbose: print(df.head())

    writer = pd.ExcelWriter('test.xlsx', engine='openpyxl')
    tday = datetime.datetime.today()
    if os.path.exists('test.xlsx') == True:
        book = load_workbook('test.xlsx')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        current = tday.strftime('%y-%m-%d-%H-%M-%S')
        print('current= ',current)
        book.create_sheet(current, 0)
    else:
        current = tday.strftime('%y-%m-%d')

    print('current= ',current)
    df.to_excel(writer, sheet_name=current)

    writer.save()


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("infile",
                        help="the input file to use to generate the Activity text")
    parser.add_argument("outfile",
                        help="the output file containing the Activity text")
    parser.add_argument("--verbose", help="increase output verbosity",
                        action="store_true")
    args = parser.parse_args()
    if args.verbose:
        print("verbosity turned on")

    do_convert(args.infile)
